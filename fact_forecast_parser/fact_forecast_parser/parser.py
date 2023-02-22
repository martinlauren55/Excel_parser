from typing import Optional
import datetime
import random
import sqlite3

import pandas as pd
from openpyxl import load_workbook


class _FactForecastBase:
    data_start_row = 4

    def __init__(self, input_data: str, output_data: str):
        """input_data - path to excel file

        output_data- path to database"""

        self.input_data = input_data
        self.output_data = output_data
        self.sheet = load_workbook(self.input_data).active

        self.length_row = self._get_row_length()

        self.db_con = sqlite3.connect(self.output_data)
        self.db_cur = self.db_con.cursor()

    def _get_row_length(self):
        len_row = self.data_start_row
        while (True):
            item = self.sheet.cell(row=len_row, column=2).value
            if item:
                len_row += 1
            else:
                len_row -= 1
                break
        return len_row

    def parse(self):
        raise NotImplementedError()

    def _create_table(self):
        raise NotImplementedError()

    def get_all_data_table(self, table_name):
        df = pd.read_sql_query(f"SELECT * FROM {table_name}", self.db_con)
        return df


class _Company(_FactForecastBase):
    company_data_start_col = 2

    def __init__(self, input_data: str, output_data: str):
        super().__init__(input_data, output_data)
        self._company_list = None
        self._parse_title()

    def _parse_title(self):
        self._title_company = self.sheet.cell(row=1, column=self.company_data_start_col).value
        self._create_table()

    def parse(self):
        """Parse file and added name of company to DB"""
        start = self.data_start_row
        company_list = []
        for i in range(start, self.length_row + 1):
            company_list.append(self.sheet.cell(row=start, column=self.company_data_start_col).value)
            start += 1
        self._company_list = set(company_list)
        self._insert_to_db()

    def _create_table(self):
        self.db_cur.execute(f"""CREATE TABLE IF NOT EXISTS {self._title_company} (
                    "id"	INTEGER,
                    "name"	TEXT NOT NULL,
                    PRIMARY KEY("id" AUTOINCREMENT)
                )""")

    def _insert_to_db(self):
        if self._company_list:
            for item in self._company_list:
                res = self.db_cur.execute(f"""SELECT name FROM {self._title_company} WHERE name='{item}'""").fetchone()
                if not res:
                    self.db_cur.execute(f"""INSERT INTO {self._title_company} (name)  VALUES ('{item}')""")
                    self.db_con.commit()
                    # self.db_cur.close()

    def get_id_by_name(self, name: str):
        """return Company id or False"""
        res = self.db_cur.execute(f"""SELECT id FROM {self._title_company} WHERE name='{name}'""").fetchone()
        if res:
            return int(res[0])
        else:
            return False


class _FactForecast(_FactForecastBase):
    def __init__(self, input_data: str, output_data: str,
                 col_num_data1: int, col_num_data2: int, table_name: str):
        super().__init__(input_data, output_data)
        self.table_name = table_name
        self._col_num_data1 = col_num_data1
        self._col_num_data2 = col_num_data2
        self._company = _Company(input_data=input_data, output_data=output_data)
        self._data_list = None

    def parse(self):
        """Parse file and added name of company to DB"""
        self._create_table()
        start = self.data_start_row
        self._data_list = []
        for i in range(start, self.length_row + 1):
            comp_name = self.sheet.cell(row=start, column=2).value
            data1 = self.sheet.cell(row=start, column=self._col_num_data1).value
            data2 = self.sheet.cell(row=start, column=self._col_num_data2).value
            self._data_list.append((comp_name, data1, data2))
            start += 1

        self._insert_to_db()

    def _create_table(self):
        self.db_cur.execute(f"""CREATE TABLE IF NOT EXISTS {self.table_name} (
                    "id"	INTEGER,
                    "company_id"	INTEGER NOT NULL,
                    "data1"	TEXT,
                    "data2"	TEXT,
                    "date" TEXT NOT NULL,
                    PRIMARY KEY("id" AUTOINCREMENT),
                    FOREIGN KEY("company_id") REFERENCES "company"("id")
                )""")

    @staticmethod
    def _get_date():
        today = datetime.date.today()
        if today.month == 2:
            return f'{today.year}-{today.month}-{random.randrange(1, 28)}'
        elif today.month in [1, 3, 4, 7, 8, 10, 12]:
            return f'{today.year}-{today.month}-{random.randrange(1, 31)}'
        else:
            return f'{today.year}-{today.month}-{random.randrange(1, 30)}'

    def _insert_to_db(self):
        if self._data_list:
            for item in self._data_list:
                comp_id = self._company.get_id_by_name(item[0])
                self.db_cur.execute(
                    f"""INSERT INTO {self.table_name} (company_id, data1, data2, date)  VALUES ({comp_id},'{item[1]}', '{item[2]}', '{self._get_date()}');""")
            self.db_con.commit()
            self.db_cur.close()


class FactForecastParser:
    def __init__(self, input_data: str, output_data: str, set_data: Optional[list] = None):
        """

        :param input_data: - Path to file Excel
        :param output_data: Path to file Database
        :param set_data: By default, the column names (table name in the database) and the column number for data1 and
                        data2 are defined. When scaling the parser file, you need to override the default value by
                        replacing or adding new columns.
        """
        self.input_data = input_data
        self.output_data = output_data
        self._set_data = set_data
        if not self._set_data:
            self._set_data = [{"table_name": "fact_Qliq", "col_num_data1": 3, "col_num_data2": 4},
                              {"table_name": "fact_Qoil", "col_num_data1": 5, "col_num_data2": 6},
                              {"table_name": "forecast_Qliq", "col_num_data1": 7, "col_num_data2": 8},
                              {"table_name": "forecast_Qoil", "col_num_data1": 9, "col_num_data2": 10}]

    def _parse_company(self):
        self._company = _Company(input_data=self.input_data, output_data=self.output_data)
        self._company.parse()

    def parse(self):
        """Run parse file"""
        self._parse_company()
        self._parse_fact_forcast()

    def _parse_fact_forcast(self):
        if self._set_data:
            for item in self._set_data:
                _FactForecast(input_data=self.input_data,
                              output_data=self.output_data,
                              col_num_data1=item['col_num_data1'],
                              col_num_data2=item['col_num_data2'],
                              table_name=item['table_name']).parse()

    def get_total(self):
        """Returns the resulting total for all tables to the console"""
        if self._set_data:
            for item in self._set_data:
                df = _FactForecastBase(input_data=self.input_data,
                                       output_data=self.output_data).get_all_data_table(item['table_name'])
                df["date"] = pd.to_datetime(df["date"])
                df["date"] = df["date"].dt.date
                df["data1"] = pd.to_numeric(df["data1"], downcast='float')
                df["data2"] = pd.to_numeric(df["data2"], downcast='float')
                df["total"] = df[["data1", "data2"]].sum(axis=1)
                print(f"\nTotal for \"{item['table_name']}\"")
                for i in range(len(df)):
                    print(df.iloc[i]['date'], df.iloc[i]['total'])
